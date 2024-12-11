import requests
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook

# ===== 設定セクション =====
EXCEL_FILE_NAME = "基本設計書_1204.xlsx"  # 保存するExcelファイル名
SHEET_NAME = "weather Data"  # 保存するシート名
REGION_XML_URL = "https://www.data.jma.go.jp/developer/xml/forecast_area.xml"  # 気象庁の府県天気予報データURL

def fetch_weather_data():
    """
    気象庁APIから天気予報データを取得
    """
    try:
        response = requests.get(REGION_XML_URL)
        response.raise_for_status()  # エラーがある場合は例外を発生
        return response.content
    except requests.exceptions.RequestException as e:
        print(f"天気データの取得に失敗しました: {e}")
        return None

def parse_weather_data(xml_data):
    """
    XML形式の天気予報データを解析して晴れや曇りの情報を抽出
    """
    weather_info = []
    root = ET.fromstring(xml_data)

    # 各エリアの予報を抽出
    for area in root.findall(".//area"):
        area_name = area.find("name").text  # 地域名
        weathers = area.findall(".//forecast")
        
        for forecast in weathers:
            date = forecast.get("date")  # 日付
            weather = forecast.find("weather").text  # 天気
            weather_info.append((area_name, date, weather))
    
    return weather_info

def save_to_excel(weather_data):
    """
    取得した天気情報をExcelファイルに保存
    """
    try:
        # 既存のExcelファイルを開く（なければ新規作成）
        try:
            workbook = load_workbook(EXCEL_FILE_NAME)
        except FileNotFoundError:
            workbook = Workbook()

        # 指定したシートが存在しなければ作成
        if SHEET_NAME not in workbook.sheetnames:
            workbook.create_sheet(title=SHEET_NAME)
        
        sheet = workbook[SHEET_NAME]

        # シートが空ならヘッダー行を追加
        if sheet.max_row == 1 and sheet.max_column == 1:
            sheet.append(["Area", "Date", "Weather"])  # カラムのヘッダー

        # データをシートに書き込む
        for area_name, date, weather in weather_data:
            sheet.append([area_name, date, weather])
        
        # Excelファイルを保存
        workbook.save(EXCEL_FILE_NAME)
        print(f"天気データを'{EXCEL_FILE_NAME}'の「{SHEET_NAME}」シートに保存しました！")

    except Exception as e:
        print(f"Excelファイルの保存中にエラーが発生しました: {e}")

def main():
    """
    メイン処理
    """
    print("気象庁APIから天気予報データを取得中...")
    xml_data = fetch_weather_data()

    if xml_data:
        print("天気予報データの解析中...")
        weather_data = parse_weather_data(xml_data)
        print("天気予報データをExcelに保存中...")
        save_to_excel(weather_data)
    else:
        print("天気予報データの取得に失敗しました。")

if __name__ == "__main__":
    main()
